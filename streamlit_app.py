import io
import re
import time
import warnings
import requests
import pandas as pd
import streamlit as st
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

warnings.filterwarnings("ignore")

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(page_title="SEC 10-K Builder", page_icon="📊", layout="wide")
st.title("📊 SEC 10-K Financial Model Builder")
st.caption("Pulls as-reported financials from SEC EDGAR and exports a formatted Excel model.")

# ── Sidebar inputs ────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("Settings")
    TICKER     = st.text_input("Ticker Symbol", value="AAPL").strip().upper()
    USER_AGENT = st.text_input("SEC User-Agent Email", value="your@email.com").strip()
    run_btn    = st.button("Build Model", type="primary", use_container_width=True)
    st.markdown("---")
    st.caption("SEC fair-use policy requires a valid email as User-Agent.")

TODAY = datetime.today().strftime("%Y-%m-%d")

# ── Style / format constants ──────────────────────────────────────────────────
NAVY       = "1F3864"
MED_GRAY   = "BFBFBF"
LIGHT_GRAY = "F2F2F2"
SUBTOTAL_C = "E8E8E8"
ALT_ROW    = "FAFAFA"
PROJ_FILL  = "EBF3FB"
WACC_INPUT = "FFF2CC"
WACC_RSLT  = "D9EAD3"
THIN_TOP   = Border(top=Side(style="thin"))
FMT_DOLLAR = '$#,##0.0;($#,##0.0);"-"'
FMT_PCT    = '0.0%;(0.0%);"-"'
FMT_EPS    = '$#,##0.00;($#,##0.00);"-"'
FMT_SHARES = '#,##0.0;(#,##0.0);"-"'
FMT_2DP    = "#,##0.00"

# ── HTTP helpers ──────────────────────────────────────────────────────────────
def _headers(url, user_agent):
    if "data.sec.gov" in url:
        return {"User-Agent": user_agent, "Accept-Encoding": "gzip, deflate", "Host": "data.sec.gov"}
    return {"User-Agent": user_agent, "Accept-Encoding": "gzip, deflate", "Host": "www.sec.gov"}

def fetch_json(url, user_agent, retries=3):
    h = _headers(url, user_agent)
    for attempt in range(retries):
        try:
            time.sleep(0.15)
            r = requests.get(url, headers=h, timeout=30)
            r.raise_for_status()
            return r.json()
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(2 ** attempt)
            else:
                raise e

def fetch_html(url, user_agent):
    h = _headers(url, user_agent)
    time.sleep(0.2)
    try:
        r = requests.get(url, headers=h, timeout=30)
        return r.text if r.status_code == 200 else None
    except Exception:
        return None

# ── XBRL helpers ──────────────────────────────────────────────────────────────
def to_millions(value, unit):
    if value is None:
        return None
    return round(value / 1_000_000, 2) if unit in ("USD", "shares") else value

def get_annual_values(facts_gaap, tag, fy_end_set):
    if tag not in facts_gaap:
        return {}
    units_dict = facts_gaap[tag].get("units", {})
    unit_key = next((k for k in units_dict if k in ("USD", "shares", "pure")), None)
    if unit_key is None:
        return {}
    annual = [e for e in units_dict[unit_key] if e.get("form") in ("10-K", "10-K/A")]
    fy_map = {}
    for e in annual:
        end = e.get("end")
        filed = e.get("filed", "")
        if end not in fy_end_set:
            continue
        if end not in fy_map or filed > fy_map[end]["filed"]:
            fy_map[end] = {"val": e.get("val"), "unit": unit_key, "filed": filed}
    return {end: to_millions(info["val"], info["unit"]) for end, info in fy_map.items()}

# ── Statement harvester ───────────────────────────────────────────────────────
_SKIP_NAMES = {"parenthetical", "supplemental", "note ", "other comprehensive",
               "geographic", "quarterly", "segment"}

_PRECISE = {
    "IS":  ["statement of income", "statements of income", "statement of operations",
            "statements of operations", "statement of earnings", "statements of earnings",
            "statements of comprehensive income"],
    "BS":  ["balance sheet", "statements of condition", "financial position", "consolidated balance"],
    "CFS": ["cash flow", "statements of cash", "cash flows"],
}
_BROAD = {
    "IS":  ["income", "operations", "earnings"],
    "BS":  ["balance"],
    "CFS": ["cash"],
}
_CAL_TAGS = {
    "IS":  ["Revenues", "RevenueFromContractWithCustomerExcludingAssessedTax", "InterestAndDividendIncomeOperating"],
    "BS":  ["Assets"],
    "CFS": ["NetCashProvidedByUsedInOperatingActivities"],
}

def _metalinks(cik, acc, user_agent):
    a = acc.replace("-", "")
    try:
        ml = fetch_json(f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{a}/MetaLinks.json", user_agent)
    except Exception:
        return None
    if ml is None:
        return None
    instance = ml.get("instance", {})
    for key, val in instance.items():
        if isinstance(val, dict) and "report" in val:
            return val
    if "report" in ml:
        return ml
    return ml

def _find_r(metalinks, stype):
    if not metalinks:
        return None
    precise = _PRECISE[stype]
    broad   = _BROAD[stype]
    reports = metalinks.get("report", {})
    best    = (0, None)
    for rk, meta in reports.items():
        ln = meta.get("longName", "").lower()
        sn = meta.get("shortName", "").lower()
        mc = meta.get("menuCat", "").lower()
        nm = f"{ln} {sn} {mc}"
        if any(w in nm for w in _SKIP_NAMES):
            continue
        score  = sum(2 for k in precise if k in nm)
        score += sum(1 for k in broad   if k in nm)
        if meta.get("groupType", "").lower() == "statement":
            score += 0.5
        if score > best[0]:
            try:
                best = (score, int(rk[1:]))
            except (ValueError, IndexError):
                pass
    return best[1]

def _scale_from_text(html):
    t = (html or "").lower()
    if re.search(r"in\s+billions",  t): return 1_000.0
    if re.search(r"in\s+millions",  t): return 1.0
    if re.search(r"in\s+thousands", t): return 0.001
    if re.search(r"\$000s?\b",     t): return 0.001
    return None

def _calibrate_scale(parsed_rows, facts_gaap, fy_end, stype):
    fy_set = {fy_end}
    cf_mm  = None
    for tag in _CAL_TAGS.get(stype, []):
        m = get_annual_values(facts_gaap, tag, fy_set)
        if m.get(fy_end) is not None:
            cf_mm = abs(m[fy_end])
            break
    if not cf_mm or cf_mm < 0.01:
        return 1.0
    for _, _, vs in parsed_rows:
        s = vs.replace(",", "").replace("$", "").strip()
        if s.startswith("(") and s.endswith(")"):
            s = s[1:-1]
        try:
            raw = abs(float(s))
        except (ValueError, TypeError):
            continue
        if raw < 1:
            continue
        ratio = cf_mm / raw
        for scale in (1.0, 0.001, 1_000.0, 1e-6, 1e6):
            if 0.70 < ratio / scale < 1.43:
                return scale
    return 1.0

def _is_num(text):
    s = (text.replace(",", "").replace("$", "").replace("(", "")
             .replace(")", "").replace("—", "").replace("–", "")
             .replace("-", "").strip())
    return s.isdigit() and len(s) > 0

def _parse_r(html, fy_end):
    if not html:
        return []
    soup  = BeautifulSoup(html, "html.parser")
    table = soup.find("table")
    if not table:
        return []
    rows = table.find_all("tr")
    if len(rows) < 2:
        return []
    col_counts = {}
    for row in rows[1:13]:
        cells = row.find_all(["td", "th"])
        for ci in range(1, len(cells)):
            if _is_num(cells[ci].get_text(strip=True)):
                col_counts[ci] = col_counts.get(ci, 0) + 1
    candidates = sorted(ci for ci, cnt in col_counts.items() if cnt >= 2)
    col_idx    = candidates[0] if candidates else 1
    data_start = 1
    for ri, row in enumerate(rows[1:6], 1):
        cells = row.find_all(["td", "th"])
        if len(cells) > col_idx and _is_num(cells[col_idx].get_text(strip=True)):
            data_start = ri
            break
    result = []
    for row in rows[data_start:]:
        cells = row.find_all(["td", "th"])
        if len(cells) <= col_idx:
            continue
        lc    = cells[0]
        label = re.sub(r"\s+", " ", re.sub(r"\[\d+\]", "", lc.get_text(" ", strip=True))).strip()
        if not label:
            continue
        sm     = re.search(r"padding-left:\s*([\d.]+)em", lc.get("style", ""))
        indent = round(float(sm.group(1))) if sm else 0
        val_str = cells[col_idx].get_text(strip=True)
        result.append((label, indent, val_str))
    return result

def _parse_val(vs, scale):
    if not vs or vs in ("—", "–", "-", "", "N/A"):
        return None
    s   = vs.replace(",", "").replace("$", "").strip()
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    try:
        return round(float(s) * scale * (-1 if neg else 1), 3)
    except (ValueError, TypeError):
        return None

def _build(raw_by_fy, fy_labels):
    order, levels, seen = [], {}, set()
    for fy in reversed(fy_labels):
        for lbl, ind, _ in raw_by_fy.get(fy, []):
            if lbl not in seen:
                order.append(lbl)
                levels[lbl] = ind
                seen.add(lbl)
    data = {}
    for lbl in order:
        row = []
        for fy in fy_labels:
            ym = {r[0]: r[2] for r in raw_by_fy.get(fy, [])}
            row.append(ym.get(lbl))
        data[lbl] = row
    df = pd.DataFrame(data, index=fy_labels).T
    df.index.name = "Line Item"
    return df, levels

# ── Excel builder (same as notebook) ─────────────────────────────────────────
_SUBTOTAL_RE = re.compile(
    r"^\s*total\b|\btotal\s|^\s*net\s+income|^\s*net\s+loss"
    r"|^\s*gross\s+profit|^\s*income\s+before\b|^\s*earnings\s+before\b"
    r"|^\s*net\s+cash\s+(provided|used)|^\s*operating\s+income"
    r"|^\s*net\s+interest\s+income|^\s*net\s+revenue|^\s*total\s+net",
    re.I)

def _auto_subtotals(df):
    return {lbl for lbl in df.index if _SUBTOTAL_RE.search(lbl)}

def _auto_section_hdrs(df, levels):
    result = set()
    for lbl in df.index:
        if lbl.strip().endswith(":"):
            result.add(lbl); continue
        if levels.get(lbl, 0) == 0:
            vals = [v for v in df.loc[lbl] if v is not None and not (isinstance(v, float) and pd.isna(v))]
            if not vals:
                result.add(lbl)
    return result

def _num_fmt(label):
    l = label.lower()
    if ("per share" in l or "per common share" in l) and "weighted" not in l:
        return FMT_EPS
    if "weighted average" in l and ("share" in l or "unit" in l):
        return FMT_SHARES
    return FMT_DOLLAR

def _row(row_map, *candidates):
    for c in candidates:
        v = row_map.get(c)
        if v is not None:
            return v
    return None

def col(n): return get_column_letter(n)

def build_excel(company_name, ticker, cik, fy_labels, proj_labels,
                df_is, is_levels, df_bs, bs_levels, df_cfs, cfs_levels):
    n_fy   = len(fy_labels)
    n_proj = len(proj_labels)
    dsc    = 3
    dec    = dsc + n_fy - 1
    psc    = dec + 1

    def set_header(ws, stmt_name):
        last_col = col(psc + n_proj - 1)
        ws.row_dimensions[1].height = 8
        ws.merge_cells(f"B2:{last_col}2")
        c = ws["B2"]
        c.value     = f"{company_name}  |  {stmt_name}"
        c.font      = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
        c.fill      = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 24
        ws["B3"].value = "All figures in $MM unless noted"
        ws["B3"].font  = Font(italic=True, size=9, color="808080", name="Calibri")
        ws.row_dimensions[3].height = 14
        hdr_fill = PatternFill("solid", fgColor=MED_GRAY)
        hdr_font = Font(bold=True, size=10, name="Calibri")
        for i, fy in enumerate(fy_labels):
            c = ws.cell(row=4, column=dsc+i, value=fy)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center")
        proj_fill = PatternFill("solid", fgColor="D9D9D9")
        for i, pl in enumerate(proj_labels):
            c = ws.cell(row=4, column=psc+i, value=pl)
            c.font = Font(bold=True, size=10, name="Calibri", color="595959")
            c.fill = proj_fill; c.alignment = Alignment(horizontal="center")

    def set_col_widths(ws):
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 42
        for i in range(n_fy):
            ws.column_dimensions[col(dsc+i)].width = 13
        for i in range(n_proj):
            ws.column_dimensions[col(psc+i)].width = 13

    def write_data_rows(ws, df, start_row, subtotals, section_headers=None, levels=None):
        sh   = section_headers or set()
        lvls = levels or {}
        r    = start_row; alt = False; rmap = {}
        for label in df.index:
            is_hdr = label in sh
            is_sub = (label in subtotals) and not is_hdr
            if is_hdr:
                indent_txt = "  " * lvls.get(label, 0) + label
                lc = ws.cell(row=r, column=2, value=indent_txt)
                lc.font = Font(bold=True, size=10, name="Calibri")
                for cn in range(2, psc + n_proj + 1):
                    ws.cell(row=r, column=cn).fill = PatternFill("solid", fgColor=LIGHT_GRAY)
                ws.row_dimensions[r].height = 16
                rmap[label] = r; r += 1; alt = False; continue
            bg    = SUBTOTAL_C if is_sub else (ALT_ROW if alt else "FFFFFF")
            rfill = PatternFill("solid", fgColor=bg)
            nfmt  = _num_fmt(label)
            indent_txt = "  " * lvls.get(label, 0) + label
            lc = ws.cell(row=r, column=2, value=indent_txt)
            lc.font = Font(bold=is_sub, size=10, name="Calibri"); lc.fill = rfill
            for i, fy in enumerate(fy_labels):
                v  = df.loc[label, fy]
                if isinstance(v, pd.Series):
                    v = v.iloc[0]
                dc = ws.cell(row=r, column=dsc+i,
                             value=(float(v) if v is not None and pd.notna(v) else None))
                dc.number_format = nfmt
                dc.font      = Font(color="0000FF", name="Calibri", size=10, bold=is_sub)
                dc.fill      = rfill
                dc.alignment = Alignment(horizontal="right")
                if is_sub: dc.border = THIN_TOP
            pfill = PatternFill("solid", fgColor=PROJ_FILL)
            for i in range(n_proj):
                pc = ws.cell(row=r, column=psc+i)
                pc.number_format = nfmt
                pc.font  = Font(color="000000", name="Calibri", size=10)
                pc.fill  = pfill
                pc.alignment = Alignment(horizontal="right")
            rmap[label] = r; ws.row_dimensions[r].height = 15
            r += 1; alt = not alt
        return r, rmap

    def write_section_hdr(ws, r, label, fg=LIGHT_GRAY):
        ws.cell(row=r, column=2, value=label).font = Font(bold=True, size=10, name="Calibri")
        for cn in range(2, psc + n_proj + 1):
            ws.cell(row=r, column=cn).fill = PatternFill("solid", fgColor=fg)
        ws.row_dimensions[r].height = 16

    def write_formula_row(ws, r, label, fn, num_fmt=FMT_PCT):
        ws.cell(row=r, column=2, value=label).font = Font(size=10, name="Calibri")
        for i in range(n_fy):
            v = fn(r, dsc+i)
            if v is None: continue
            dc = ws.cell(row=r, column=dsc+i, value=v)
            dc.number_format = num_fmt
            dc.font      = Font(color="000000", name="Calibri", size=10)
            dc.alignment = Alignment(horizontal="right")
        ws.row_dimensions[r].height = 15

    def write_stats_block(ws, start_r, row_map, stat_labels, fmt_map=None):
        write_section_hdr(ws, start_r, "SUMMARY STATISTICS", fg="DEE6EF")
        hdr_r = start_r + 1
        for cn, txt in [(2, "Line Item"), (3, "5-Yr CAGR"),
                        (4, f"5-Yr Avg  ({fy_labels[0]}–{fy_labels[-1]})")]:
            c = ws.cell(row=hdr_r, column=cn, value=txt)
            c.font = Font(bold=True, size=9, name="Calibri", color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="2E5F8A")
            c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[hdr_r].height = 14
        c_s = col(dsc); c_e = col(dec); n_p = n_fy - 1 if n_fy > 1 else 1
        r   = hdr_r + 1; alt = False
        for label in stat_labels:
            dr = row_map.get(label)
            if dr is None: continue
            fill    = PatternFill("solid", fgColor=ALT_ROW if alt else "FFFFFF")
            num_fmt = (fmt_map or {}).get(label, FMT_DOLLAR)
            ws.cell(row=r, column=2, value=label).font = Font(size=9, name="Calibri")
            ws.cell(row=r, column=2).fill = fill
            cc = ws.cell(row=r, column=3,
                         value=f"=IFERROR(({c_e}{dr}/{c_s}{dr})^(1/{n_p})-1,\"-\")")
            cc.number_format = FMT_PCT; cc.font = Font(bold=True, size=9, name="Calibri")
            cc.fill = fill; cc.alignment = Alignment(horizontal="center")
            ac = ws.cell(row=r, column=4,
                         value=f"=IFERROR(AVERAGE({c_s}{dr}:{c_e}{dr}),\"-\")")
            ac.number_format = num_fmt; ac.font = Font(bold=True, size=9, name="Calibri")
            ac.fill = fill; ac.alignment = Alignment(horizontal="right")
            ws.row_dimensions[r].height = 14; r += 1; alt = not alt
        return r

    wb = Workbook(); wb.remove(wb.active)

    # Cover
    cover = wb.create_sheet("Cover")
    cover.column_dimensions["A"].width = 4
    cover.column_dimensions["B"].width = 30
    cover.column_dimensions["C"].width = 44
    cover.merge_cells("B2:D2")
    t = cover["B2"]
    t.value     = f"{company_name}  |  SEC 10-K Financial Model"
    t.font      = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
    t.fill      = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center")
    cover.row_dimensions[2].height = 30
    most_recent_year = int(fy_labels[-1][2:])
    for i, (k, v) in enumerate([
        ("Company",       company_name),
        ("Ticker",        ticker.upper()),
        ("CIK",           cik),
        ("Date Pulled",   TODAY),
        ("Fiscal Years",  ", ".join(fy_labels)),
        ("Forecast Cols", ", ".join(proj_labels)),
        ("Units",         "All values in $MM unless per-share"),
        ("Data Source",   "SEC EDGAR XBRL Viewer R-files (as reported)"),
    ], start=4):
        cover.cell(row=i, column=2, value=k).font = Font(bold=True, name="Calibri", size=10)
        cover.cell(row=i, column=3, value=v).font = Font(name="Calibri", size=10)
    cover.cell(row=13, column=2, value="Contents").font = Font(bold=True, size=11, name="Calibri")
    for i, sname in enumerate(["Income Statement", "Balance Sheet",
                                "Cash Flow Statement", "WACC", "As-Reported Labels"], start=14):
        c = cover.cell(row=i, column=2, value=sname)
        c.hyperlink = f"#{sname}!A1"
        c.font = Font(color="0563C1", underline="single", name="Calibri", size=10)

    # Income Statement
    ws_is = wb.create_sheet("Income Statement")
    set_col_widths(ws_is); set_header(ws_is, "Income Statement")
    sub_is  = _auto_subtotals(df_is)
    shdr_is = _auto_section_hdrs(df_is, is_levels)
    next_r, is_map = write_data_rows(ws_is, df_is, 5, sub_is, shdr_is, is_levels)
    next_r += 1; write_section_hdr(ws_is, next_r, "MARGINS & RATIOS"); next_r += 1
    rv = _row(is_map, "Total net interest and noninterest income", "Total revenues",
              "Net revenue", "Revenues", "Revenue", "Total interest and noninterest income")
    gp = _row(is_map, "Gross profit", "Gross Profit")
    oi = _row(is_map, "Operating income", "Income from operations",
              "Income before income taxes", "Operating Income")
    ni = _row(is_map, "Net income attributable to Popular, Inc.", "Net income",
              "Net Income", "Net income attributable to common shareholders")
    tx = _row(is_map, "Income tax expense", "Income taxes",
              "Provision for income taxes", "Income Tax Expense")
    pt = _row(is_map, "Income before income taxes", "Pretax income", "Pretax Income")
    def mfn(nr, dr):
        if nr and dr:
            return lambda r, c: f"=IFERROR({col(c)}{nr}/{col(c)}{dr},\"-\")"
        return lambda r, c: None
    for lbl, fn in [
        ("Gross Margin %",       mfn(gp, rv)),
        ("Operating Margin %",   mfn(oi, rv)),
        ("Net Margin %",         mfn(ni, rv)),
        ("Effective Tax Rate %", mfn(tx, pt)),
    ]:
        write_formula_row(ws_is, next_r, lbl, fn, FMT_PCT); next_r += 1
    if rv:
        ws_is.cell(row=next_r, column=2, value="YoY Revenue Growth %").font = Font(size=10, name="Calibri")
        for i in range(n_fy):
            if i == 0: continue
            dc = ws_is.cell(row=next_r, column=dsc+i,
                            value=f"=IFERROR({col(dsc+i)}{rv}/{col(dsc+i-1)}{rv}-1,\"-\")")
            dc.number_format = FMT_PCT
            dc.font = Font(color="000000", name="Calibri", size=10)
            dc.alignment = Alignment(horizontal="right")
        ws_is.row_dimensions[next_r].height = 15; next_r += 1
    next_r += 1
    eps_lbl = next((l for l in is_map if "diluted" in l.lower() and "per share" in l.lower()), None)
    stat_is_rows = [rv, _row(is_map, "Net income attributable to Popular, Inc.", "Net income", "Net Income")]
    stat_labels_is = [lbl for lbl, rn in is_map.items()
                      if rn in [x for x in stat_is_rows if x] or lbl in {eps_lbl}]
    write_stats_block(ws_is, next_r, is_map, stat_labels_is, {eps_lbl: FMT_EPS} if eps_lbl else {})

    # Balance Sheet
    ws_bs = wb.create_sheet("Balance Sheet")
    set_col_widths(ws_bs); set_header(ws_bs, "Balance Sheet")
    sub_bs  = _auto_subtotals(df_bs)
    shdr_bs = _auto_section_hdrs(df_bs, bs_levels)
    next_r_bs, bs_map = write_data_rows(ws_bs, df_bs, 5, sub_bs, shdr_bs, bs_levels)
    next_r_bs += 1; write_section_hdr(ws_bs, next_r_bs, "KEY RATIOS"); next_r_bs += 1
    ca  = _row(bs_map, "Total current assets", "Total Current Assets", "Total assets")
    cl  = _row(bs_map, "Total current liabilities", "Total Current Liabilities")
    te  = _row(bs_map, "Total stockholders' equity", "Total equity", "Total Equity", "Total shareholders' equity")
    ltd = _row(bs_map, "Long-term debt", "Long-term borrowings", "Long-Term Debt")
    csh = _row(bs_map, "Cash and cash equivalents", "Cash & Equivalents", "Cash and due from banks")
    ta  = _row(bs_map, "Total assets", "Total Assets")
    for lbl, fn, fmt in [
        ("Current Ratio",
         (lambda r, c: f"=IFERROR({col(c)}{ca}/{col(c)}{cl},\"-\")") if ca and cl else (lambda r,c: None),
         FMT_2DP),
        ("Total Debt / Total Assets",
         (lambda r, c: f"=IFERROR({col(c)}{ltd}/{col(c)}{ta},\"-\")") if ltd and ta else (lambda r,c: None),
         FMT_2DP),
        ("Net Debt ($MM)",
         (lambda r, c: f"=IFERROR({col(c)}{ltd}-{col(c)}{csh},\"-\")") if ltd and csh else (lambda r,c: None),
         FMT_DOLLAR),
    ]:
        write_formula_row(ws_bs, next_r_bs, lbl, fn, fmt); next_r_bs += 1
    next_r_bs += 1
    bs_stat_lbls = [lbl for lbl in bs_map
                    if any(k in lbl.lower() for k in ("total assets", "total liabilities",
                                                       "total equity", "total deposits"))]
    write_stats_block(ws_bs, next_r_bs, bs_map, bs_stat_lbls)

    # Cash Flow
    ws_cfs = wb.create_sheet("Cash Flow Statement")
    set_col_widths(ws_cfs); set_header(ws_cfs, "Cash Flow Statement")
    sub_cfs  = _auto_subtotals(df_cfs)
    shdr_cfs = _auto_section_hdrs(df_cfs, cfs_levels)
    next_r_cfs, cfs_map = write_data_rows(ws_cfs, df_cfs, 5, sub_cfs, shdr_cfs, cfs_levels)
    next_r_cfs += 1; write_section_hdr(ws_cfs, next_r_cfs, "KEY METRICS"); next_r_cfs += 1
    ocf_r = _row(cfs_map,
        "Net cash provided by operating activities",
        "Net cash provided by (used in) operating activities",
        "Net Cash Provided by Operating Activities")
    cpx_r = _row(cfs_map,
        "Purchases of premises and equipment", "Capital expenditures",
        "Purchase of property and equipment", "Payments to acquire property plant and equipment")
    ni_c  = _row(cfs_map, "Net income", "Net income attributable to Popular, Inc.", "Net Income (CFS)")
    for lbl, fn, fmt in [
        ("CapEx % of Total Assets",
         (lambda r, c: f"=IFERROR(ABS({col(c)}{cpx_r})/'Balance Sheet'!{col(c)}{ta},\"-\")") if cpx_r and ta else (lambda r,c: None),
         FMT_PCT),
        ("OCF / Net Income",
         (lambda r, c: f"=IFERROR({col(c)}{ocf_r}/{col(c)}{ni_c},\"-\")") if ocf_r and ni_c else (lambda r,c: None),
         FMT_2DP),
    ]:
        write_formula_row(ws_cfs, next_r_cfs, lbl, fn, fmt); next_r_cfs += 1
    next_r_cfs += 1
    cfs_stat_lbls = [lbl for lbl in cfs_map
                     if any(k in lbl.lower() for k in ("operating activities", "investing activities",
                                                        "financing activities", "net change in cash"))]
    write_stats_block(ws_cfs, next_r_cfs, cfs_map, cfs_stat_lbls)

    # WACC
    ws_w = wb.create_sheet("WACC")
    for c_, w in [("A",4),("B",38),("C",16),("D",16),("E",13),("F",13),("G",13)]:
        ws_w.column_dimensions[c_].width = w
    def w_title(r, text):
        ws_w.merge_cells(f"B{r}:G{r}")
        c = ws_w[f"B{r}"]
        c.value = text; c.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
        c.fill  = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws_w.row_dimensions[r].height = 26
    def w_sec(r, text):
        ws_w.merge_cells(f"B{r}:G{r}")
        c = ws_w[f"B{r}"]
        c.value = text; c.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        c.fill  = PatternFill("solid", fgColor="2E5F8A")
        ws_w.row_dimensions[r].height = 16
    def w_row(r, label, value=None, formula=None, fmt=FMT_PCT, is_input=False, is_result=False):
        lc = ws_w.cell(row=r, column=2, value=label)
        lc.font = Font(bold=is_result, size=10, name="Calibri")
        if is_result: lc.fill = PatternFill("solid", fgColor=WACC_RSLT)
        vc = ws_w.cell(row=r, column=3, value=formula if formula else value)
        vc.number_format = fmt; vc.alignment = Alignment(horizontal="center")
        if is_input:
            vc.fill = PatternFill("solid", fgColor=WACC_INPUT)
            vc.font = Font(color="0000FF", size=10, name="Calibri")
        elif is_result:
            vc.fill = PatternFill("solid", fgColor=WACC_RSLT)
            vc.font = Font(bold=True, color="375623", size=11, name="Calibri")
        else:
            vc.font = Font(color="000000", size=10, name="Calibri")
        ws_w.row_dimensions[r].height = 15
    w_title(2, f"{company_name}  |  WACC Buildup")
    ws_w["B3"].value = "Yellow = user inputs.  Green = WACC result."
    ws_w["B3"].font  = Font(italic=True, size=9, color="808080", name="Calibri")
    w_sec(5, "COST OF DEBT")
    w_row(6,  "Pre-tax cost of debt  (Kd)",               value=0.045,  fmt=FMT_PCT, is_input=True)
    w_row(7,  "Marginal tax rate  (T)",                    value=0.21,   fmt=FMT_PCT, is_input=True)
    w_row(8,  "After-tax cost of debt  Kd × (1-T)",       formula="=C6*(1-C7)", fmt=FMT_PCT)
    w_sec(10, "COST OF EQUITY  —  CAPM")
    w_row(11, "Risk-free rate  (Rf)  10-yr UST",           value=0.043,  fmt=FMT_PCT, is_input=True)
    w_row(12, "Equity beta  (β)",                          value=1.20,   fmt=FMT_2DP, is_input=True)
    w_row(13, "Equity risk premium  (ERP)",                value=0.055,  fmt=FMT_PCT, is_input=True)
    w_row(14, "Cost of equity  Ke = Rf + β × ERP",        formula="=C11+C12*C13", fmt=FMT_PCT)
    w_sec(16, "CAPITAL STRUCTURE")
    for cn, txt in [(2,"Component"),(3,"Market Value ($MM)"),(4,"Target Override"),(5,"Weight")]:
        c = ws_w.cell(row=17, column=cn, value=txt)
        c.font = Font(bold=True, size=9, name="Calibri")
        c.fill = PatternFill("solid", fgColor=MED_GRAY)
        c.alignment = Alignment(horizontal="center")
    ws_w.row_dimensions[17].height = 14
    for r_, lbl, we_formula in [
        (18, "Equity (Market Capitalization)",      "=IF(D18<>\"\",D18,C18/(C18+C19))"),
        (19, "Net Debt  (Total Debt − Cash)",       "=1-E18"),
    ]:
        ws_w.cell(row=r_, column=2, value=lbl).font = Font(size=10, name="Calibri")
        inp = ws_w.cell(row=r_, column=3)
        inp.fill = PatternFill("solid", fgColor=WACC_INPUT)
        inp.font = Font(color="0000FF", size=10, name="Calibri")
        inp.number_format = FMT_DOLLAR; inp.alignment = Alignment(horizontal="center")
        we = ws_w.cell(row=r_, column=5, value=we_formula)
        we.number_format = FMT_PCT; we.font = Font(color="000000", size=10, name="Calibri")
        we.alignment = Alignment(horizontal="center"); ws_w.row_dimensions[r_].height = 15
    w_sec(21, "WACC")
    ws_w.cell(row=22, column=2, value="WACC  =  Ke × We  +  Kd(1−T) × Wd").font = Font(bold=True, size=11, name="Calibri")
    ws_w.cell(row=22, column=2).fill = PatternFill("solid", fgColor=WACC_RSLT)
    wr = ws_w.cell(row=22, column=3, value="=C14*E18+C8*E19")
    wr.number_format = FMT_PCT; wr.font = Font(bold=True, size=13, color="375623", name="Calibri")
    wr.fill = PatternFill("solid", fgColor=WACC_RSLT)
    wr.alignment = Alignment(horizontal="center", vertical="center")
    ws_w.row_dimensions[22].height = 22

    # As-Reported Labels
    ws_raw = wb.create_sheet("As-Reported Labels")
    ws_raw.freeze_panes = "A2"
    ws_raw.append(["Statement", "Label (as filed)", "Indent (em)"] + fy_labels)
    raw_rows = []
    for stmt, df_map, lvl_map in [
        ("Income Statement",    df_is,  is_levels),
        ("Balance Sheet",       df_bs,  bs_levels),
        ("Cash Flow Statement", df_cfs, cfs_levels),
    ]:
        for label in df_map.index:
            row = [stmt, label, lvl_map.get(label, 0)]
            for fy in fy_labels:
                v = df_map.loc[label, fy]
                row.append(float(v) if pd.notna(v) and v is not None else None)
            ws_raw.append(row); raw_rows.append(row)
    n_raw = len(raw_rows) + 1
    tbl   = Table(displayName="AsReportedLabels", ref=f"A1:{col(3+n_fy)}{n_raw}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws_raw.add_table(tbl)
    ws_raw.column_dimensions["A"].width = 20
    ws_raw.column_dimensions["B"].width = 55
    ws_raw.column_dimensions["C"].width = 12
    for i in range(n_fy):
        ws_raw.column_dimensions[col(4+i)].width = 13

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── Main pipeline ─────────────────────────────────────────────────────────────
def run_pipeline(ticker, user_agent):
    status = st.status("Running pipeline…", expanded=True)

    with status:
        st.write("Resolving ticker → CIK…")
        tickers_data = fetch_json("https://www.sec.gov/files/company_tickers.json", user_agent)
        cik_raw = company_name = None
        for entry in tickers_data.values():
            if entry["ticker"].upper() == ticker.upper():
                cik_raw      = entry["cik_str"]
                company_name = entry["title"]
                break
        if cik_raw is None:
            st.error(f"Ticker '{ticker}' not found in SEC ticker list.")
            return None
        CIK = str(cik_raw).zfill(10)
        st.write(f"✅ {ticker} → {company_name} (CIK: {CIK})")

        st.write("Fetching 10-K filing list…")
        submissions = fetch_json(f"https://data.sec.gov/submissions/CIK{CIK}.json", user_agent)
        filings     = submissions.get("filings", {}).get("recent", {})
        tenk_filings = [
            {"accessionNumber": filings["accessionNumber"][i],
             "filingDate":      filings["filingDate"][i],
             "reportDate":      filings["reportDate"][i],
             "primaryDocument": filings["primaryDocument"][i]}
            for i, form in enumerate(filings.get("form", []))
            if form in ("10-K", "10-K/A")
        ]
        tenk_filings.sort(key=lambda x: x["filingDate"], reverse=True)
        tenk_filings = tenk_filings[:5]
        tenk_filings.sort(key=lambda x: x["reportDate"])
        FY_ENDS   = [f["reportDate"] for f in tenk_filings]
        FY_LABELS = [f"FY{f['reportDate'][:4]}" for f in tenk_filings]
        MOST_RECENT_YEAR = int(FY_ENDS[-1][:4])
        PROJ_LABELS = [f"FY{MOST_RECENT_YEAR + i}E" for i in range(1, 4)]
        st.write(f"✅ {len(tenk_filings)} filings: {', '.join(FY_LABELS)} | Forecast: {', '.join(PROJ_LABELS)}")

        st.write("Downloading XBRL company facts…")
        facts_data = fetch_json(f"https://data.sec.gov/api/xbrl/companyfacts/CIK{CIK}.json", user_agent)
        facts_gaap = facts_data.get("facts", {}).get("us-gaap", {})
        st.write(f"✅ {len(facts_gaap)} us-gaap concepts")

        st.write("Parsing as-reported statements from SEC XBRL Viewer…")
        _raw = {"IS": {}, "BS": {}, "CFS": {}}
        for filing in tenk_filings:
            acc    = filing["accessionNumber"]
            fy_e   = filing["reportDate"]
            fy_lbl = f"FY{fy_e[:4]}"
            ml     = _metalinks(CIK, acc, user_agent)
            if not ml:
                st.warning(f"{fy_lbl}: MetaLinks not found — skipping")
                for st_ in _raw: _raw[st_][fy_lbl] = []
                continue
            for stype in ("IS", "BS", "CFS"):
                rn = _find_r(ml, stype)
                if rn is None:
                    _raw[stype][fy_lbl] = []
                    continue
                a    = acc.replace("-", "")
                html = fetch_html(
                    f"https://www.sec.gov/Archives/edgar/data/{int(CIK)}/{a}/R{rn}.htm",
                    user_agent)
                rows  = _parse_r(html, fy_e)
                scale = _scale_from_text(html) or _calibrate_scale(rows, facts_gaap, fy_e, stype)
                final = [(lbl, ind, _parse_val(vs, scale)) for lbl, ind, vs in rows]
                _raw[stype][fy_lbl] = final
                st.write(f"  ✅ {fy_lbl} {stype} → R{rn:02d}  ({len(final)} rows, scale ×{scale:.4g})")

        df_is,  is_levels  = _build(_raw["IS"],  FY_LABELS)
        df_bs,  bs_levels  = _build(_raw["BS"],  FY_LABELS)
        df_cfs, cfs_levels = _build(_raw["CFS"], FY_LABELS)

        st.write("Building Excel workbook…")
        excel_buf = build_excel(company_name, ticker, CIK, FY_LABELS, PROJ_LABELS,
                                df_is, is_levels, df_bs, bs_levels, df_cfs, cfs_levels)

    status.update(label="Done!", state="complete", expanded=False)
    return {
        "company_name": company_name,
        "ticker":       ticker.upper(),
        "cik":          CIK,
        "fy_labels":    FY_LABELS,
        "proj_labels":  PROJ_LABELS,
        "df_is":        df_is,
        "df_bs":        df_bs,
        "df_cfs":       df_cfs,
        "excel_buf":    excel_buf,
    }

# ── Run & display ─────────────────────────────────────────────────────────────
if run_btn:
    if not USER_AGENT or "@" not in USER_AGENT:
        st.error("Please enter a valid email address as SEC User-Agent.")
    else:
        st.session_state["result"] = run_pipeline(TICKER, USER_AGENT)

result = st.session_state.get("result")

if result:
    r = result
    st.subheader(f"{r['company_name']} ({r['ticker']})")
    col1, col2, col3 = st.columns(3)
    col1.metric("CIK", r["cik"])
    col2.metric("Fiscal Years", ", ".join(r["fy_labels"]))
    col3.metric("Forecast", ", ".join(r["proj_labels"]))

    fname = f"{r['ticker']}_10K_{TODAY}.xlsx"
    st.download_button(
        label="⬇️  Download Excel Model",
        data=r["excel_buf"],
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )

    tab_is, tab_bs, tab_cfs = st.tabs(["Income Statement", "Balance Sheet", "Cash Flow Statement"])

    def display_df(tab, df, label):
        with tab:
            st.caption(f"{len(df)} line items  |  all figures in $MM")
            fmt = {c: "{:,.1f}" for c in df.columns}
            st.dataframe(df.style.format(fmt, na_rep="—"), use_container_width=True, height=600)

    display_df(tab_is,  r["df_is"],  "Income Statement")
    display_df(tab_bs,  r["df_bs"],  "Balance Sheet")
    display_df(tab_cfs, r["df_cfs"], "Cash Flow Statement")
